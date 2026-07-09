from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

MISSING_SITES_EXPORT_FILENAME = "clientes_con_sedes_sin_diagrama.xlsx"

MISSING_SITES_COLUMNS = ("cliente", "sede", "calle", "provincia")
MISSING_SITES_HEADERS = {
    "cliente": "Cliente",
    "sede": "Sede",
    "calle": "Calle",
    "provincia": "Provincia",
}


_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r", "\n")


def _sanitize_cell(value):
    """Prevent Excel/CSV formula injection from external (GLPI) data."""
    text = "" if value is None else str(value)
    if text and text[0] in _FORMULA_TRIGGERS:
        return "'" + text
    return text


def missing_sites_to_xlsx(rows: list[dict]) -> bytes:
    """Build an Excel workbook listing sedes without a published diagram."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sin diagrama"

    header_font = Font(bold=True)
    for column_index, key in enumerate(MISSING_SITES_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=MISSING_SITES_HEADERS[key])
        cell.font = header_font

    max_lengths = {key: len(MISSING_SITES_HEADERS[key]) for key in MISSING_SITES_COLUMNS}
    for row_index, row in enumerate(rows, start=2):
        for column_index, key in enumerate(MISSING_SITES_COLUMNS, start=1):
            raw = row.get(key, "")
            max_lengths[key] = max(max_lengths[key], len(str(raw)))
            sheet.cell(row=row_index, column=column_index, value=_sanitize_cell(raw))

    for column_index, key in enumerate(MISSING_SITES_COLUMNS, start=1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = min(max(max_lengths[key] + 2, 12), 60)

    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
